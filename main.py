import os
import sys
import tqdm
from nicegui import app, ui, events
import logging
import asyncio
from io import StringIO
from logging import getLogger, StreamHandler
from pydub import AudioSegment
import math
import glob
import openpyxl
import simpleaudio
from moviepy import VideoFileClip

os.environ['PYWEBVIEW_GUI'] = 'qt' # needed when running on ubuntu

''' global arguments '''
audio_segment_max_size = 25000000 # bytes (whisper max: ~25 000 000)
audio_segment_overlap_ms = 2000  # overlap all segments by 2s if possible
models = [
    'tiny', 'base', 'small', 'medium', 'large', 'turbo'
]
languages = [
    'Auto',
    'Afrikaans', 'Arabic', 'Armenian', 'Azerbaijani', 'Belarusian', 'Bosnian', 'Bulgarian', 'Catalan', 'Chinese', 'Croatian', 'Czech', 'Danish', 'Dutch',
    'English', 'Estonian', 'Finnish', 'French', 'Galician', 'German', 'Greek', 'Hebrew', 'Hindi', 'Hungarian', 'Icelandic', 'Indonesian', 'Italian',
    'Japanese', 'Kannada', 'Kazakh', 'Korean', 'Latvian', 'Lithuanian', 'Macedonian', 'Malay', 'Marathi', 'Maori', 'Nepali', 'Norwegian', 'Persian',
    'Polish', 'Portuguese', 'Romanian', 'Russian', 'Serbian', 'Slovak', 'Slovenian', 'Spanish', 'Swahili', 'Swedish', 'Tagalog', 'Tamil', 'Thai', 
    'Turkish', 'Ukrainian', 'Urdu', 'Vietnamese', 'Welsh'
]
output_formats = [
    'xlsx', 'srt', 'txt', 'vtt', 'tsv', 'json'
]

class LogElementHandler(logging.Handler):
    """A logging handler that emits messages to a log element."""

    def __init__(self, element: ui.log, level: int = logging.NOTSET) -> None:
        self.element = element
        super().__init__(level)

    def emit(self, record: logging.LogRecord) -> None:
        try:
            msg = self.format(record)
            self.element.push(msg)
        except Exception:
            self.handleError(record)

logger = logging.getLogger()

class ViewModel:
    """ holds all texts and values to be displayed by the ui """
    def __init__(self):
        self.button_file_content = 'choose audio files'
        self.selected_files = None
        self.label_progress_content = ''
        self.button_run_enabled = False
        self.spinner_progress_visibility = False
        self.file_count = 0
        self.file_count_old = 0
        self.segment_count = 0
        self.segment_done_count = 0
        self.segment_current_progress = 0
        self.selected_output_formats = []
            
    def update_label_progress(self):
        if self.file_count <= 0:
            self.spinner_progress_visibility = False
            self.label_progress_content = ''
            self.segment_current_progress = 0
            self.segment_count = 0
            self.segment_done_count = 0
            ui.notify('finished transcribing')
            ViewModel.play_sound_effect_finished()
        else:
            if self.file_count_old == 0:
                self.spinner_progress_visibility = True
            info = ''
            if self.file_count == 1:
                info += f'transcribing {self.file_count} file'
            else:
                info += f'transcribing {self.file_count} files'
            if self.segment_count == 1:
                info += f' ({self.segment_done_count} / {self.segment_count} segment done)'
            else:
                info += f' ({self.segment_done_count} / {self.segment_count} segments done)'
            if self.segment_current_progress > 0:
                info += f'\ncurrent segment {self.segment_current_progress}% done'
            self.label_progress_content = info
        self.file_count_old = self.file_count
    
    def update_buttons(self):
        # check whether any files are selected
        if self.selected_files is None or len(self.selected_files) == 0:
            self.button_run_enabled = False
            self.button_file_content = 'choose audio / video files'
        else:
            if len(self.selected_files) == 1:
                self.button_file_content = '1 file selected'
            else:    
                self.button_file_content = f'{len(self.selected_files)} files selected'
            # check whether any output_formats are selected
            if self.selected_output_formats is None or len(self.selected_output_formats) == 0:
                self.button_run_enabled = False
            else:
                self.button_run_enabled = True

    def get_output_language(language) -> str:
        language_return = None
        if language is not None and language != 'Auto':
            language_return = language  
        return language_return
        
    def update_select_output_formats(self, e: events.ValueChangeEventArguments):
        self.selected_output_formats = e.value
        self.update_buttons()

    def toggle_mute():
        app.storage.general['mute'] = not app.storage.general['mute']

    def play_sound_effect_finished():
        if (not app.storage.general['mute']):
            sound_effect_path = "sound_effect_finished.wav"
            if not os.path.isfile(sound_effect_path):
                sound_effect_path = os.path.join("_internal", "sound_effect_finished.wav")
            if sound_effect_path is not None:
                wave_obj = simpleaudio.WaveObject.from_wave_file(sound_effect_path)
                wave_obj.play()

viewmodel = ViewModel()

class CustomProgressBar(tqdm.tqdm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._current = self.n  # Set the initial value
        
    def update(self, n):
        super().update(n)
        self._current += n
        global viewmodel
        # Inform listeners
        viewmodel.segment_current_progress = round(self._current / self.total * 100)
        viewmodel.update_label_progress()

import whisper.transcribe
transcribe_module = sys.modules['whisper.transcribe']
transcribe_module.tqdm.tqdm = CustomProgressBar # inject progressbar into tqdm.tqdm of whisper, so we can see progress
import whisper

async def start_reading_console():
    """ start a 'stream' of console outputs """
    global viewmodel
    string_io = StringIO() # Create buffer
    sys.stdout = string_io # Standard output like a print
    sys.stderr = string_io # Errors/Exceptions
    stream_handler = StreamHandler(string_io) # Logmessages
    stream_handler.setLevel("DEBUG")
    logger.addHandler(stream_handler)
    while 1:
        await asyncio.sleep(1)  # need to update ui
        logger.info(string_io.getvalue())
        string_io.truncate(0)

logger = getLogger(__name__)
logger.setLevel("DEBUG")

extensions_audio = ('.mp3','.m4a','.m4b','.m4p','.flac','.ogg','.oga','.mogg','.wav','.wma','.mmf','.aa','.aax')
filepicker_formats_audio = "Audio Files (*.mp3;*.m4a;*.m4b;*.m4p;*.flac;*.ogg;*.oga;*.mogg;*.wav;*.wma;*.mmf;*.aa;*.aax)"
extensions_video = ('.webm','.mkv','.flv','.vob','.ogv','.ogg','.drc','.avi','.mts','.m2ts','.ts','.mov','.qt','.wmv','.rm','.rmvb','.viv','.asf','.amv','.mp4','.m4p','.m4v','.mpg','.mp2','.mpeg','.mpe','.mpv','.m2v','.m4v','.svi','.3gp','.3g2','.f4v','.f4p','.f4a','.f4b')
filepicker_formats_video = "Video Files (*.webm;*.mkv;*.flv;*.vob;*.ogv;*.ogg;*.drc;*.avi;*.mts;*.m2ts;*.ts;*.mov;*.qt;*.wmv;*.rm;*.rmvb;*.viv;*.asf;*.amv;*.mp4;*.m4p;*.m4v;*.mpg;*.mp2;*.mpeg;*.mpe;*.mpv;*.m2v;*.m4v;*.svi;*.3gp;*.3g2;*.f4v;*.f4p;*.f4a;*.f4b)"
async def choose_files():
    """ open a file picker to select multiple files """
    global viewmodel
    viewmodel.selected_files = await app.native.main_window.create_file_dialog(allow_multiple=True, file_types=["All Files (*)", filepicker_formats_audio, filepicker_formats_video])
    #check whether any files need to be split    
    need_splitting_count = 0
    if viewmodel.selected_files != None:
        for file in viewmodel.selected_files:
            file_size = os.path.getsize(file) #size in bytes
            if file_size > audio_segment_max_size:
                need_splitting_count += 1
                if need_splitting_count > 1:
                    break
        if need_splitting_count > 0:
            if need_splitting_count == 1:
                ui.notify('1 file is too large and needs to be split')
            else:    
                ui.notify(f'{need_splitting_count} files are too large and need to be split')
    viewmodel.update_buttons()

class AudioSplitter:
    def split_audio(file: str) -> list[str, int]:
        """ 
        checks the size of an audio file and breaks it into equal smaller files if it is larger than <audio_segment_max_size> 
        
        it returns a list containing the filepath to the split audio segments and their start timestamp in milliseconds
        """
        split_files = []
        segment_count = AudioSplitter.get_segment_count(file)
        if segment_count > 1:
            temp_dir = AudioSplitter.create_temp_dir()
            song = AudioSegment.from_file(file)
            print(f'\nsplitting {file} into {segment_count} parts')
            segment_length_ms = len(song) / segment_count #len(song): length of song in ms (might differ from compressed audiofile as pydub converts to wav when loading audio file)
            overlap_ms = audio_segment_overlap_ms
            if segment_length_ms < audio_segment_overlap_ms:
                overlap_ms = 0
            for i in range (segment_count):
                seg_start_ms = math.floor(i * segment_length_ms - overlap_ms / 2)
                if i == 0:
                    seg_start_ms = None
                seg_end_ms = math.ceil((i + 1) * segment_length_ms + overlap_ms / 2)
                if i == segment_count - 1:
                    seg_end_ms = None
                segment = song[seg_start_ms:seg_end_ms]
                segment_filename = os.path.join(temp_dir, os.path.splitext(os.path.basename(file))[0] + '_segment' + str(i) + os.path.splitext(file)[1])
                segment.export(segment_filename)
                print(f'\n  > saved segment {segment_filename}')
                split_files.append((segment_filename, seg_start_ms))
            print(f'\nfinished splitting {file}')
        else:
            split_files.append((file, 0))
        return split_files
        
    def get_segment_count(file : str) -> int:
        """ returns the number of segments the specified file needs to be split, in order for each segment not to exceed audio_segment_max_size """
        file_size = os.path.getsize(file) #size in bytes
        if file_size > audio_segment_max_size:
            return math.ceil(file_size / audio_segment_max_size)
        else:
            return 1

    def get_temp_dir() -> str:
        if getattr(sys, 'frozen', False):
            return os.path.join(os.path.dirname(sys.executable), '.temp')
        elif __file__:
            return os.path.join(os.path.dirname(__file__), '.temp')
        return 'C:\\_lokal\\.temp'

    def create_temp_dir() -> str:
        temp_dir = AudioSplitter.get_temp_dir()
        print(f'\ntemp dir: {temp_dir}')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
        return temp_dir

    def clear_temp_dir():
        """ removes all files in <temp_dir> """
        temp_dir = AudioSplitter.get_temp_dir()
        files = glob.glob(os.path.join(temp_dir,'*'))
        for f in files:
            os.remove(f)

class AudioExtractor:
    def file_is_video(file: str) -> bool:
        return file.endswith(extensions_video)

    def extract_audio_from_video(file: str) -> str:
        """
        returns the file_path of the audio output
        """
        # create the filepath of the audiofile
        output_audio_path = os.path.splitext(file)[0] + '.mp3'
        try:
            # Load the video file
            video = VideoFileClip(file)        
            # Extract audio
            audio = video.audio
            # Save the audio file
            audio.write_audiofile(output_audio_path)
        except Exception:
            print(Exception)
            print(f'could not extract audio from movie > {file}')
            return None
        
        return output_audio_path

def whisper_transcribe(files: list[str], model_str: str, language_str: str , output_format: list[str]):
    global transcribe_module, viewmodel
    print(f'\nloading model {model_str}, this might take some time ...')
    transcribe_module = whisper.load_model(model_str)
    for file in files:
        if AudioExtractor.file_is_video(file):
            audio_file = AudioExtractor.extract_audio_from_video(file)
            if audio_file is not None:
                file = audio_file
        file_segments = AudioSplitter.split_audio(file)
        results = []
        if file_segments != None and len(file_segments) > 0:
            for i in range (len(file_segments)):
                results.append(transcribe_module.transcribe(file_segments[i][0], language=language_str, verbose=True, fp16=False))
                viewmodel.segment_done_count += 1
                viewmodel.update_label_progress()
        # # DEBUGGING: export each segment individually
        # for i in range(len(results)):
        #     whisper_save_result(results[i], ['txt'], os.path.splitext(file_segments[i][0])[0]+'.txt')
        # combine all segments into one and export it
        if len(results) > 1:
            # append all text to the first result
            # start at second segment
            for i in range (1, len(results)):
                results[0]['text'] += '\n' + results[i]['text']
                # also add all segments to the first result and fix their timestamps
                for segment in results[i]["segments"]:
                    segment["start"] += file_segments[i][1] / 1000 # ms to s
                    segment["end"] += file_segments[i][1] / 1000 # ms to s
                    results[0]['segments'].append(segment)
            AudioSplitter.clear_temp_dir()
        whisper_save_result(results[0], output_format, file)

def whisper_save_result(result, output_ext: list[str], file_path: str):
    output_dir = os.path.dirname(os.path.realpath(file_path))
    for ext in output_ext:
        output_filename = os.path.splitext(file_path)[0] + '.' + ext
        if ext == 'xlsx':
            wb = openpyxl.Workbook() 
            sheet = wb.active
            sheet.cell(row=1, column=1).value = 'start' # frikin excel starting rows/columns at 1   (╯°□°）╯︵ ┻━┻
            sheet.cell(row=1, column=2).value = 'end'
            sheet.cell(row=1, column=3).value = 'text'
            for i in range(len(result["segments"])):
                sheet.cell(row=i+2, column=1).value = whisper.utils.format_timestamp(result["segments"][i]["start"])
                sheet.cell(row=i+2, column=2).value = whisper.utils.format_timestamp(result["segments"][i]["end"])
                sheet.cell(row=i+2, column=3).value = result["segments"][i]["text"]
            wb.save(os.path.join(output_dir, output_filename))

            # with open(os.path.join(output_dir, output_filename), 'w', newline='', encoding='UTF8') as csvfile:
            #     csv_writer = csv.writer(csvfile, delimiter=';')
            #     csv_writer.writerow(['start','end','text'])
            #     for segment in result["segments"]:
            #         csv_writer.writerow([whisper.utils.format_timestamp(segment["start"]), whisper.utils.format_timestamp(segment["end"]), segment["text"]])

        else:
            output_writer = whisper.utils.get_writer(ext, output_dir)
            output_writer(result, output_filename)

async def start_transcribing(files, model, language, output_format):
    global viewmodel
    viewmodel.file_count += len(files)
    viewmodel.selected_files = None
    viewmodel.update_buttons()
    # get total number of segments that will be processed
    segcount = 0
    for file in files:
        segcount += AudioSplitter.get_segment_count(file)
    viewmodel.segment_count += segcount
    viewmodel.update_label_progress()
    # start transcribing
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, lambda: whisper_transcribe(files, model, ViewModel.get_output_language(language), output_format))
    # update labels after transcribing ends
    viewmodel.file_count -= len(files)
    viewmodel.segment_count -= segcount
    viewmodel.update_label_progress()

@ui.page('/')
def main():
    """ contains all nicegui elements which make up the interface """
    global viewmodel

    # initialize ui states in storage.general
    if not 'selected_model' in app.storage.general:
        app.storage.general['selected_model'] = 'tiny'
    if not 'selected_language' in app.storage.general:
        app.storage.general['selected_language'] = 'German'
    if not 'selected_output_format' in app.storage.general:
        app.storage.general['selected_output_format'] = ['xlsx', 'txt']
    if not 'dark_mode' in app.storage.general:
        app.storage.general['dark_mode'] = None
    if not 'mute' in app.storage.general:
        app.storage.general['mute'] = False
    dark_mode = ui.dark_mode().bind_value(app.storage.general, 'dark_mode')
    # build page
    with ui.column().classes('w-full'):
        with ui.row().classes('w-full items-center'):
            ui.icon('record_voice_over', color='primary').classes('text-4xl')
            ui.label('Whisper Transcribe').classes('text-primary').style('font-size: 150%')
            ui.space()
            with ui.column():
                ui.button(icon='auto_mode', on_click=dark_mode.disable) \
                    .props('outline round').tooltip('automatic theme').bind_visibility_from(dark_mode, 'value', lambda mode: mode is None)
                ui.button(icon='light_mode', on_click=dark_mode.enable) \
                    .props('outline round').tooltip('light theme').bind_visibility_from(dark_mode, 'value', value=False)
                ui.button(icon='dark_mode', on_click=dark_mode.auto) \
                    .props('outline round').tooltip('dark theme').bind_visibility_from(dark_mode, 'value', value=True)
            with ui.column():
                ui.button(icon='volume_up', on_click=ViewModel.toggle_mute) \
                    .props('outline round').tooltip('play sound').bind_visibility_from(app.storage.general, 'mute', value=False)
                ui.button(icon='volume_off', on_click=ViewModel.toggle_mute) \
                    .props('outline round').tooltip('mute').bind_visibility_from(app.storage.general, 'mute', value=True)
        ui.button(icon='insert_drive_file', on_click=choose_files).bind_text_from(viewmodel, 'button_file_content').style('margin-top: 8px')
        ui.select(options=models, label='model').classes('w-full').bind_value(app.storage.general, 'selected_model')
        ui.select(options=languages, label='language', with_input=True).classes('w-full').bind_value(app.storage.general, 'selected_language')
        ui.select(options=output_formats, label='output', multiple=True, on_change=viewmodel.update_select_output_formats).classes('w-full').bind_value(app.storage.general, 'selected_output_format').props('use-chips')
        ui.label('Results are saved in the same directory as the original files.').style('color: #808080; font-style: italic; margin-top: 16px')
        ui.button('start', icon='auto_awesome', on_click=lambda: start_transcribing(viewmodel.selected_files, app.storage.general['selected_model'], app.storage.general['selected_language'], app.storage.general['selected_output_format'])).bind_enabled_from(viewmodel, 'button_run_enabled')
        with ui.row().classes('w-full justify-center'):
            ui.spinner('dots', size='xl').bind_visibility_from(viewmodel, 'spinner_progress_visibility')
        ui.label().classes('w-full text-center').style('color: #808080; font-style: italic; white-space: pre-wrap').bind_text_from(viewmodel, 'label_progress_content')
        with ui.expansion().classes('w-full') as expansion:
            ui.query('.nicegui-expansion .q-expansion-item__content').style('padding:0', replace='gap:0')
            with expansion.add_slot('header'):
                with ui.row().classes('w-full items-center'):
                    ui.label('console output').style('color: #808080')
                    ui.space()
            ui_log = ui.log(max_lines=100).classes("w-full h-40").style('white-space: pre-wrap')
            handler = LogElementHandler(ui_log)
            logger.addHandler(handler)
            ui.context.client.on_disconnect(lambda: logger.removeHandler(handler))
    
def main():
    app.on_startup(start_reading_console)
    ui.run(title='Whisper Transcribe', reload=False, native=True, window_size=[500,800], storage_secret='foobar')

if __name__ == '__main__':
    main()

